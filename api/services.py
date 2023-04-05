import requests
from tqdm import tqdm
from openpyxl import load_workbook
import config
from materials import materials
from composite_layers import composite_layers


def start_render_job(attributes, layers, asset_id, stage_id):
    url = "https://preview.threekit.com/api/asset-jobs/renders"

    payload = {
        "settings": {
            "output": {
                "resolution": {
                    "width": 3010
                }
            }
        },
        "orgId": f"{config.ORG_ID}",
        "assets": [
            {
                "assetId": f"{asset_id}",
                "subset": {
                    "attributes": attributes,
                    "layers": {
                        "filterMode": "include",
                        "values": layers
                    },
                    "uniqueAttrs": "all"
                }
            }
        ],
        "stages": [
            {
                "assetId": f"{stage_id}",
                "subset": {
                    "attributes": {
                        "Camera Position": {
                            "values": [
                                "0",
                                "1",
                                "2",
                                "3",
                                "4"
                            ],
                            "filterMode": "include"
                        }
                    },
                    "layers": {
                        "filterMode": "include"
                    },
                    "uniqueAttrs": "all"
                }
            }
        ],
        "jobName": "",
        "type": "vray",
        "renderMissingLayers": False
    }
    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "authorization": "Bearer 2fcb22e4-c5b3-4d76-8db1-fbcce0b326f7"
    }

    print('starting job')

    response = requests.post(url, json=payload, headers=headers).json()

    print(response)

    return response['jobId']


def get_materials(attributes):
    material_values = {}
    for attribute in attributes:
        print('\n\n')
        print(attribute['name'])
        temp = []
        for value in attribute["values"]:
            try:
                temp.append(materials[value["label"]])
                print(f'- {value["label"]}')
            except KeyError:
                # print(value["label"], " not found in materials")
                pass
        material_values[attribute["name"]] = {"values": temp, "filterMode": "include"}
    return material_values


def get_layers(attributes, composite_id):
    layers = []

    attribute_names = [attribute['name'] for attribute in attributes]

    layers.append(composite_layers[composite_id]['B'])
    layers.append(composite_layers[composite_id]['P1'])
    layers.append(composite_layers[composite_id]['P2'])
    layers.append(composite_layers[composite_id]['P3'])

    return layers


def parse_xlsx(filename):
    sku_list = []

    workbook = load_workbook(filename=filename)
    worksheet = workbook.active

    for row in tqdm(worksheet.iter_rows(min_row=2, min_col=2, values_only=True), desc=f'Reading {filename}'):
        sku_list.append(row[0])

    return sku_list


def sku_to_id(sku):
    """ids_list = []

    for sku in tqdm(sku_list, desc='SKU -> asset_id'):
        url = f'https://threekitapi.herokuapp.com/vray-sku/{sku}/'
        response = requests.get(url)
        asset_id = response.text
        if asset_id:
            ids_list.append(asset_id)

    return ids_list"""

    url = f'https://threekitapi.herokuapp.com/vray-sku/{sku}/'
    response = requests.get(url)
    asset_id = response.text

    return asset_id


def get_composite(asset_id):
    url = f'https://preview.threekit.com/api/v2/assets/{asset_id}'

    headers = {
        "accept": "application/json",
        "authorization": f"Bearer {config.BEARER_TOKEN}"
    }

    response = requests.get(url, headers=headers).json()

    return response['defaultCompositeId']


def get_stage_id(asset_id):
    url = f'https://preview.threekit.com/api/v2/assets/{asset_id}'

    headers = {
        "accept": "application/json",
        "authorization": f"Bearer {config.BEARER_TOKEN}"
    }

    response = requests.get(url, headers=headers).json()

    return response['defaultStageId']